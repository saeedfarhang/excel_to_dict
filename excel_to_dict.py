from openpyxl import load_workbook

def excel_to_dict(file):

    items_list = []
    item_dict = {}

    wb = load_workbook(filename=file)
    # get the first work sheet
    ws = wb.worksheets[0]
    # get the number of rows
    row_count = ws.max_row

    # keys are the values of first row
    keys = []
    for i in ws[1]:
        keys.append(i.value)

    for row in range(row_count-1):
        values = []
        for i in ws.columns:
            values.append(i[row+1].value)
        
        for key in keys:
            item_dict[key] = values[keys.index(key)]
        items_list.append(item_dict)
    
    return items_list
